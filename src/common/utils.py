import csv
import os.path
from copy import copy
from decimal import Decimal

from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.styles.numbers import FORMAT_CURRENCY_EUR_SIMPLE
from openpyxl.utils import get_column_letter
from typing import Optional, Union

Numeric = Union[Decimal, float, int]
MatrixElement = Union[Numeric, str, datetime, None]
Matrix = [[MatrixElement]]


def write_csv(contents: [[str]], full_path_to_file: str) -> None:
    with open(full_path_to_file, 'w') as f:
        csv_writer = csv.writer(f)
        for line in contents:
            if line is not None:
                csv_writer.writerow(line)


# TODO: Write a wrapper around openpyxl, that represents an Excel file, including its path to file to save it
def write_xlsx(contents: [[str]], full_path_to_file: str, workbook: Workbook = None,
               worksheet_name: str = None) -> Workbook:
    if not worksheet_name:
        worksheet_name = extract_name_from_full_path_to_file(full_path_to_file)

    # If the full path to the file does not have an extension, we add the default extension.
    # If it has an extension however, we keep whatever the client passes.
    if os.path.splitext(full_path_to_file)[-1] == '':
        full_path_to_file += '.xlsx'

    if not workbook:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = worksheet_name
    else:
        worksheet = workbook.create_sheet(worksheet_name)

    # Set the column widths proportional to the content length. Works only approximately for nontrue type fonts.
    column_widths = [0] * len(contents[0])
    for line in contents:
        worksheet.append(line)
        for i, cell in enumerate(line):
            cell_width = len(str(cell))
            if cell_width > column_widths[i]:
                column_widths[i] = cell_width
    for i, column_width in enumerate(column_widths):
        worksheet.column_dimensions[get_column_letter(i + 1)].width = column_width

    # Set currency. Must be set after the column widths, otherwise they disappear
    # TODO: With the file wrapper, make helper functions for these kinds of operations
    for cell in worksheet['C0'][1:]:
        cell.number_format = '"€"#,##0.00'
    for cell in worksheet['F0'][1:]:
        cell.number_format = '"€"#,##0.00'

    # Highlight the header
    for i in range(len(column_widths)):
        worksheet['{}1'.format(get_column_letter(i + 1))].fill = PatternFill('solid', fgColor='BBBBBB')

    # Highlight the totals row
    default_font = worksheet['A1'].font
    bold_font = copy(default_font)
    bold_font.bold = True
    for i in range(len(column_widths)):
        worksheet['{}{}'.format(get_column_letter(i + 1), len(contents))].fill = PatternFill('solid', fgColor='DDDDDD')
        worksheet['{}{}'.format(get_column_letter(i + 1), len(contents))].font = bold_font

    workbook.save(full_path_to_file)
    return workbook


def extract_name_from_full_path_to_file(full_path_to_file: str) -> str:
    """
    Extract a sane name from a given filename

    :param full_path_to_file: Full path to a file
    :return: The base of the filename, stripped from extension and path information
    """

    filename = os.path.split(full_path_to_file)[-1]
    return os.path.splitext(filename)[0]


def concatenate_matrices(left: Matrix, right: Matrix) -> Matrix:
    """
    Put two matrices side by side

    :param left: Matrix-shaped contents
    :param right: Matrix-shaped contents
    :return: Matrix-shaped contents
    """

    # Make sure the matrices have equal amount of rows
    length_difference = len(left) - len(right)
    if length_difference > 0:
        # The left matrix is larger than the right matrix: add empty rows to the right matrix
        nr_columns = len(right[0])
        right += [[None] * nr_columns] * length_difference
    elif length_difference < 0:
        # The right matrix is larger than the left matrix: add empty rows to the left matrix
        nr_columns = len(left[0])
        left += [[None] * nr_columns] * -length_difference

    # Merge the matrices row by row
    return [left_row + right_row for left_row, right_row in zip(left, right)]


def sum_available_elements(list_of_elements: [Optional[float]]) -> Optional[float]:
    """
    Return the sum of all elements that are not None. Return None if all elements are None.
    """

    available_elements = [element for element in list_of_elements if element is not None]
    if available_elements:
        return sum(available_elements)
    else:
        return None
